from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

# Ruta del archivo Excel
EXCEL_FILE = "datos.xlsx"

# Verifica si el archivo existe, si no, lo crea
try:
    wb = load_workbook(EXCEL_FILE)
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.append(["Nombre", "Correo", "Comentario"])  # Agrega encabezados
    wb.save(EXCEL_FILE)

@app.route('/add', methods=['POST'])
def add_data():
    """Endpoint para agregar datos al archivo Excel"""
    data = request.json
    nombre = data.get("nombre")
    correo = data.get("correo")
    comentario = data.get("comentario")

    if not (nombre and correo and comentario):
        return jsonify({"error": "Todos los campos son obligatorios"}), 400

    # Abre el archivo Excel y agrega una nueva fila
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nombre, correo, comentario])
    wb.save(EXCEL_FILE)
    return jsonify({"message": "Datos agregados exitosamente"}), 201

@app.route('/get', methods=['GET'])
def get_data():
    """Endpoint para obtener todos los datos del archivo Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Lee los datos del archivo Excel
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Omite encabezados
    response = {index + 1: {"nombre": row[0], "correo": row[1], "comentario": row[2]} for index, row in enumerate(rows)}

    return jsonify(response), 200

if __name__ == '__main__':
    app.run(debug=True)
